
# -*- coding: utf-8 -*-
"""
Created on Tue Aug 20 14:32:18 2019

@author: pengdk
"""

import os
import openpyxl    #读写xlsx 
import datetime
from openpyxl.styles import Color,Font,Alignment,PatternFill,Border,Side,Protection
import shutil

import BaseData as BD

#======================================================================================================================
class Trader(object):
    def __init__(self, name):
        self._name = name
        self._codemap = {}    # key = code  value = []
        pass
    
    def addCode(self, code, rowno):
        if code in self._codemap:
            self._codemap[code].append( rowno )
            pass
        else:
            self._codemap[code] = [ rowno ]
        pass
    
    def getName(self):
        return self._name
    
    def getCodemap(self):
        return self._codemap

#=======================================================================================================================
class historyAccount(object):
    def __init__(self):
        self._tradermap = {}    #key=交易员  value=trader
        pass
    
    def setData(self, holiday_obj, snapdata_obj):
        self._oholiday = holiday_obj
        self._osnapdata = snapdata_obj
        pass

    #计算剩余天数(从今日开始，计算到enddate中间的工作日，算尾不算头)
    def calcLeftDays(self, enddate, dt_today):
        return self._oholiday.calcLeftDays(enddate, dt_today)
    
    def calcPreWorkdate(self, dt_today):
        return self._oholiday.preWorkDay(dt_today)
    
    def isValidDate(self, dt):
        if dt is None:
            return False
        if isinstance(dt, datetime.datetime):
            return True
        return False
    
    def helperdate(self, dt):
        d = dt.date()
        return d.year*10000 +  d.month*100 + d.day

    def readAccount(self, curdir, dt_today):
        filepath = os.path.join(curdir, "场外台账.xlsx")
        
        self._dt_today = dt_today
        today = dt_today.year*10000 + dt_today.month*100 + dt_today.day
        
        wb = openpyxl.load_workbook( filepath)
        ws = wb.active
        
        rows = []
        #按行循环
        idx = 0
        for row in ws.iter_rows():
            idx += 1
            if idx == 1:        #第一行是标题
                continue
            
            trader = row[0].value   #交易员
            if trader is None:      #没有交易员的行，是无效的
                continue
                
            if len(trader) ==0:
                continue
                
            enddate = self.helperdate( row[6].value )#最后到期日
            if enddate < today:
                continue
            
            later = row[19].value   #平仓日期
            if self.isValidDate(later):
                el_date = self.helperdate( later )
                if el_date < today:
                    continue
                
            pcdate = row[22].value   #结算日期
            if self.isValidDate(pcdate):
                el_date = self.helperdate( pcdate )
                if el_date < today:
                    continue
                
            #TODO
            ##因为是历史版本，台账中有未来数据,所以，任何开始日期大于today的，也要过滤    
            begindate = self.helperdate( row[5].value )   
            if begindate > today:
                continue
            #TODO    
                
            #补丁 2019-8-29
            trader.strip()
            if trader in self._tradermap:
                pass
            else:
                self._tradermap[trader] = Trader(trader)
            #end
            
            rows.append(  row )

            idx += 1
            
        print( len(rows))  
        return self.filterRead(rows)


    def filterRead(self, rows):
        t = self._dt_today
        today = t.year*10000 + t.month*100 + t.day
        
        today_rows = []
        later_rows = []
        
        '''
        当日存续的产品中，再次区分；"非当日了结" 和 "当日了结"
        当日了结:
            1. 最后日期为今日
            2. 平仓日期有值，并且等于今日
            3. 行权日期有值，并且等于今日
        '''
        
        for  row  in rows:
            enddate = self.helperdate( row[6].value )#最后到期日
            if enddate == today:
                today_rows.append(row)
                continue
            
            later = row[19].value   #平仓日期
            if self.isValidDate(later):
                el_date = self.helperdate( later )
                if el_date == today:
                    today_rows.append(row)
                    continue
            
            pcdate = row[22].value   #结算日期
            if self.isValidDate(pcdate):
                el_date = self.helperdate( pcdate )
                if el_date == today:
                    today_rows.append(row)
                    continue
                
            later_rows.append(  row )    
            pass
        
        return today_rows, later_rows

    def writeHeader(self, ws):
        titles = [
        ("交易员" , 10),                    #   A
        ("接单（主动or被动）", 10),          #   B
        ("状态",10 ),                       #   C
        ("客户名称", 30),                   #    D
        ("客户类别",	 10),               #    E
        ("开始日", 20),                     #    F
        ("到期日", 20),                     #    G
        ("标的	", 10),                     #    H
        ("客户交易方向	", 10),              #    I
        ("类型	", 10),                     #    J
        
        ("数量(吨）", 10),                  #     K           
        ("行权价", 10),                     #     L
        ("入场价", 10),	                    #     M
        ("单位权利金", 10),	                    #   N
        ("权利金", 10),	                    #     O
        ("类型", 10),                       #     P
        ("客户交易方向", 10),               #      Q
        ("进场波动度", 10),                 #       R
        ("天数	", 10),                     #       S
        ("时间	", 10),                     #       T
        
        ("到期时间	", 10),                 #       U
        ("期权计算日	", 20),             #       V
        ("昨日期权计算日	", 20),         #       W
        ("Risk-free rate ( r )	", 10),     #       X
        ("Cost of carry ( b )	", 10),     #       Y
        ("今日收盘价格	", 10),                 #   Z
        ("昨日收盘价格	", 10),             #       AA
        ("今日市价	", 10),                 #       AB
        ("昨日市价	", 10),                 #       AC
        
        ("期权dialy持仓盈亏	", 20),	        #       AD
        ("台账编号", 10),                   #       AE
        ("确认书编号", 20),                 #       AF
        ("期权今日市价估值", 20),           #       AG
        ("交易方向(我方)", 10),             #   AH
        ("期权今日浮动盈亏", 20)            #   AI
        ]
        
        cols_name = ['A', 'B', 'C', 'D', 'E','F','G','H','I','J',
              'K','L','M','N','O','P','Q','R','S','T',
              'U','V','W','X','Y','Z','AA','AB','AC','AD',
              'AE', 'AF', 'AG', 'AH', 'AI']
        
        fill_heading = PatternFill('solid', fgColor='C5D9F1')  
        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        ft = Font(name=u'宋体',size=11)
        
        no = 0
        for one in titles:
            # 调整列宽
            ws.column_dimensions[ cols_name[no] ].width = one[1]
            
            cell = ws.cell(row=1, column=no+1)
            cell.value = one[0]
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)  # 水平居中，垂直居中
            cell.fill = fill_heading
            cell.border = border
            cell.font = ft
            no += 1

            pass
        # 调整行高
        ws.row_dimensions[1].height = 40
        pass
    
        
    def writeContent(self, ws, rows,  begin_rowno):
        #cols = 15
        
        fill_content = PatternFill('solid', fgColor='D9D9D9')  # 灰色
        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        ft = Font(name=u'宋体',size=11)
        fill_xie = PatternFill('solid', fgColor='F79646') 
        fill_liu = PatternFill('solid', fgColor='DA9694') 
        fill_he = PatternFill('solid', fgColor='B8CCE4') 
        fill_lin = PatternFill('solid', fgColor='B1A0C7') 
        fill_blue = PatternFill('solid', fgColor='2FA4E7')
        fill_pink = PatternFill('solid', fgColor='F4B084') 
        
        rowno = begin_rowno
        
        for row in rows:
            cur_code = row[7].value
            _precloseprice =0
            _closeprice =0
            try:
                _precloseprice, _closeprice = self._osnapdata.getClosePrice(cur_code)
            except:
                pass
                
            for i in range(35):
                cell = ws.cell(rowno + 1, i+1)
                cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平居中，垂直居中
                cell.border = border
                cell.font = ft
                cell.fill = fill_content
                if i<14:    #前15个直接从台账中拷贝
                    cell.value= row[i].value
                    if i == 2:
                        clear_no = row[16].value
                        if clear_no is not None:
                            cell.value = "到期"
                        else: 
                            cell.value = "存续"
                        pass
                    
                    if i ==3:  #其中第4列 需要按交易员涂 背景颜色
                        trader = row[0].value
                        if trader == '谢晨星':
                            cell.fill = fill_xie
                        elif trader == '刘剑溥':
                            cell.fill = fill_liu
                        elif trader == '何剑桥':
                            cell.fill = fill_he
                        elif trader == '林秉玮':
                            cell.fill = fill_lin
                    if i == 5 or i== 6:
                        cell.number_format = 'yyyy/m/d'
                        pass
                    if i==7:
                        trader = row[0].value
                        trader.strip()
               
                        self._tradermap[trader].addCode(cur_code, rowno + 1)
                        
                        pass
                    
                elif i == 14:  #权利金
                    cell.value = "=N{0}*K{0}".format(rowno+1)
                elif i == 15:  #类型C/P
                    cell.value = 'C' if row[9].value and len(row[9].value) >2 and (row[9].value)[-2:] == '看涨' else 'P'
                elif i == 16:  #B/S
                    cell.value = 'B' if row[8].value == "买入" else 'S'
                elif i == 17:  #进场波动率
                    cell.value = row[28].value
                    cell.number_format = '0.00%'
                elif i == 18:  #天数
                    cell.value = self.calcLeftDays( row[6].value, self._dt_today )
                    pass
                elif i == 19:   #时间
                    cell.value = '15:00:01'
                elif i == 20:   #到期时间
                    cell.value = row[6].value
                    cell.number_format = 'yyyy/m/d'
                    pass
                elif i == 21:   #期权计算日
                    cell.value = self._dt_today.strftime("%Y-%m-%d ") + "15:00:01"
                    pass
                elif i == 22:   #昨日期权计算日
                    cell.value = self.calcPreWorkdate(self._dt_today).strftime("%Y-%m-%d ") + "15:00:01"
                elif i == 23:   
                    cell.value = 0.045
                    cell.number_format = '0.00%'
                elif i == 24:
                    cell.value =0
                elif i == 25:    #今天收盘价格
                    cell.value = _closeprice
                    #cell.value = r"='C:\Wind\Wind.NET.Client\WindNET\DataBrowse\XLA\WindFunc.xla'!S_DQ_Close(H{0},V{0})".format(rowno+1)
                elif i == 26:    #昨日收盘价格
                    cell.value = _precloseprice
                    #cell.value = r"='C:\Wind\Wind.NET.Client\WindNET\DataBrowse\XLA\WindFunc.xla'!S_DQ_Close(H{0},W{0})".format(rowno+1)
                elif i == 27:    #今日估值
                    cell.value = "=GBlackScholes(P{0},Z{0},L{0},S{0}/245,X{0},Y{0},R{0})".format(rowno+1)
                    #cell.number_format = '0.0000_'
                elif i == 28:    #昨日估值
                    cell.value = "=GBlackScholes(P{0},AA{0},L{0},(S{0}+1)/245,X{0},Y{0},R{0})".format(rowno+1)
                    #cell.number_format = '0.0000_'
                elif i == 29:    #持仓daily盈亏
                    cell.value = '=IF(I{0}="卖出",(AB{0}-AC{0})*K{0},(AC{0}-AB{0})*K{0})'.format(rowno+1)
                    #cell.number_format = '0.0000_'
                elif i == 30:   #台账中的编号
                    cell.value = row[27].value
                elif i == 31:   #确认书编号
                    cell.value = row[15].value
                elif i == 32:   #期权今日市价估值
                    cell.value = '=AB{}'.format(rowno + 1)
                elif i == 33:   #交易方向(我方)
                    cell.value = -1 if row[8].value == "买入" else 1
                elif i == 34:   #期权今日浮动盈亏
                    cell.value = '=(AG{0}-N{0})*K{0}*AH{0}'.format(rowno+1)
                else:
                    cell.value = ""
                pass
            rowno += 1
            ws.row_dimensions[rowno].height = 20
        
        #汇总，  盈亏汇总
        if (rowno - begin_rowno) >0:
            cell = ws.cell(rowno + 1, 29+1)
            cell.value = "=SUM(AD{0}:AD{1})".format(begin_rowno+1, rowno)
            cell.fill = fill_blue
            cell.border = border

            #20191015 add
            cell = ws.cell(rowno + 1, 34+1)
            cell.value = "=SUM(AI{0}:AI{1})".format(begin_rowno+1, rowno)
            cell.fill = fill_pink
            cell.border = border
            
        pass
    
    def reArrangeContent(self, ws, rows, begin_rowno):
        fill_yellow = PatternFill('solid', fgColor='FFFF00')  
        fill_pink = PatternFill('solid', fgColor='FCD5B4')
        
        '''
        有4种  半特殊的:
            1, 当日到期的， 今日估值为0， 昨日估值为BS
            2，当日平仓的， 今日估值为 人工填写(平仓价), 昨日估值为BS
            3，当日行权的， 今日估值为 人工填写，  昨日估值为BS
            4，开始日期为今天的， 今日估值为BS， 昨日估值为  单位权利金，即入场价
        '''
        
        t = self._dt_today
        today = t.year*10000 + t.month*100 + t.day
        
        rowno = begin_rowno
        for row in rows:
            
            begin_date = row[5].value
            end_date = row[6].value
            close_date = row[19].value
            clear_date = row[22].value
            
            rowno += 1 
            
            if self.helperdate(begin_date) == today:
                cell = ws.cell( rowno, 28+1)
                cell.value = row[13].value
                cell.fill = fill_pink
                
                ws.cell(rowno, 2+1).value = "新增"
                continue
                
            if self.isValidDate(close_date):
                if self.helperdate(close_date) == today:
                    cell = ws.cell( rowno, 27+1)
                    cell.value = row[17].value
                    cell.fill = fill_pink
                    
                    ws.cell(rowno, 2+1).value = "先平"
                    continue
                
            if self.helperdate( end_date) == today:
                cell = ws.cell( rowno, 27+1)
                #TODO v1 来自谷宇(已被替代)
                '''
                看涨期权：Max（标的期货收盘价 – 行权价，0 ）
                看跌期权：Max（行权价 – 标的期货收盘价，0 ）
                cell.value = '=IF(P{0}="C",MAX(s_dq_close(H{0},V{0})-L{0}, 0), MAX(L{0}-s_dq_close(H{0},V{0}), 0)  )'.format(rowno)
                '''
                
                #TODO v2 来自小海
                '''
                看涨期权：Max（台账中的结算价 – 行权价，0 ）
                看跌期权：Max（行权价 – 台账中的结算价，0 ）
                '''
                
                if row[20].value is not None:
                    cell.value = '=IF(P{0}="C",MAX({1}-L{0}, 0), MAX(L{0}-{1}, 0)  )'.format(rowno, row[20].value)
                else:
                    cell.value = ""
                    
                cell.fill = fill_pink
                
                ws.cell(rowno, 2+1).value = "到期"
                continue    
            
            if self.isValidDate(clear_date):
                if self.helperdate(clear_date) == today:
                    cell = ws.cell( rowno, 27+1)
                    #TODO v2 来自小海
                    '''
                    看涨期权：Max（台账中的结算价 – 行权价，0 ）
                    看跌期权：Max（行权价 – 台账中的结算价，0 ）
                    '''
                    
                    if row[20].value is not None:
                        cell.value = '=IF(P{0}="C",MAX({1}-L{0}, 0), MAX(L{0}-{1}, 0)  )'.format(rowno, row[20].value)
                    else:
                        cell.value = ""
                        
                    cell.fill = fill_pink
                    
                    ws.cell(rowno, 2+1).value = "先结"
                    continue
                
            pass
        
        '''
        有3类特殊的，要标识黄色:
            1, 标的是价差
            2, 类型不是4个字
            3, 备注：远期，互换，保险
        '''
        
        rowno = begin_rowno
        for row in rows:
            bret = False
            
            code = row[7].value         #代码,代码长度大于6
            cptype = row[9].value       #类型
            remark = row[30].value
            if len(code) >8:
                bret = True
            
            if cptype and len(cptype) != 4:
                bret = True
            
            if remark:
                if remark.find('互换') >=0:
                    bret = True
                if remark.find('远期') >=0:
                    bret = True
                if remark.find('保险') >=0:
                    bret = True
                if remark.find('Collar') >=0:
                    bret = True
                if remark.find('Delta') >=0:
                    bret = True
            
            if bret:
                for i in range(35):
                    cell = ws.cell(rowno + 1, i+1)
                    if i <30 or i == 32:
                        cell.fill = fill_yellow
                        
                    if i == 27 or i==28:    
                        cell.value = ''
                    pass
              
            rowno += 1    
            pass
        pass
    
    #写入当日盈亏 Daily盈亏
    def writeDaily(self, wb, later_rows, today_rows):
        fill_blue = PatternFill('solid', fgColor='2FA4E7')
        fill_pink = PatternFill('solid', fgColor='F4B084') 

        #按序号获取sheet
        ws = wb.worksheets[1]
        
        #[1]写入标题
        self.writeHeader(ws)
        
        #[2]写入 非当日了结 内容，并check
        begin_rowno = 1
        self.writeContent(ws, later_rows, begin_rowno)
        self.reArrangeContent(ws, later_rows, begin_rowno)
        
        #[3]写入 当日了结 内容， 并check
        begin_rowno += len(later_rows) + 5
        self.writeContent(ws, today_rows, begin_rowno)
        self.reArrangeContent(ws, today_rows, begin_rowno)
        
        #[4]写入总汇总
        rowno_1 = 2 + len(later_rows)
        rowno_2 = 5 + rowno_1 + len(today_rows)
        cell = ws.cell(rowno_2 +1, 32)
        cell.value = "=AD{0}+AD{1}".format(rowno_1,  rowno_2)
        cell.fill = fill_blue
        #20191015 add
        cell = ws.cell(rowno_2 +1, 35)
        cell.value = "=AI{0}+AI{1}".format(rowno_1,  rowno_2)
        cell.fill = fill_pink
        
        return rowno_2 +1
    
    def writeTrader( self, wb, lastrow):
        ws = wb.worksheets[1]
        
        begin_row = lastrow + 5
        ws.cell( begin_row +1, 5).value = "盈亏"
        
        rowno = begin_row+2
        for _,obj in self._tradermap.items():
            ws.cell(rowno, 1).value = obj.getName()
            codemap = obj.getCodemap()
            for k, v in codemap.items():
                ws.cell(rowno, 2).value = k
                ws.cell(rowno, 5).value = "= " + "+".join( ["AD{0}".format(i) for i in v ] )
                
                rowno += 1
            pass
            
        pass
    
    def writeTemplate(self, filepath, later_rows, today_rows):
        
        wb = openpyxl.load_workbook(filepath, keep_vba = True)
    
        #写入当日场外的盈亏
        lastrow = self.writeDaily( wb, later_rows,  today_rows )
        
        #写入按交易员的统计
        self.writeTrader( wb, lastrow)
            
        # Save the file
        wb.save(filepath)
        
        return None


#=======================================================================================================================
def main(rootdir, dt_today):
    reportdir = os.path.join( rootdir, "report")
    
    filename = "{0}-{1}-{2}损益日报表(收盘价).xlsm".format( dt_today.year,  dt_today.month, dt_today.day)
    filepath = os.path.join(reportdir, filename)
    
    datadir = os.path.join( rootdir, "dataotc")
    srcpath = os.path.join(  datadir, "Blank模板勿删(收盘价).xlsm")
    if os.path.exists(filepath):
        pass
    else:
        shutil.copyfile(srcpath,   filepath)
        
    #[1]
    oholiday = BD.CHolidayData()
    pre_workday = oholiday.preWorkDay(dt_today)

    baseobj = BD.CSnapData4(dt_today,  pre_workday)
    baseobj.start()

    #[2]
    obj = historyAccount()
    obj.setData(oholiday, baseobj)
    today_rows, later_rows = obj.readAccount(datadir,  dt_today)
    obj.writeTemplate(filepath, later_rows, today_rows)
    
    pass





   
    