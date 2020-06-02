# -*- coding: utf-8 -*-
"""
Created on Tue Aug 20 10:25:47 2019

@author: pengdk
"""

import re
import os
import datetime
import codecs 
import pandas as pd
import openpyxl
from openpyxl.styles import Color,Font,Alignment,PatternFill,Border,Side,Protection
import shutil

import  WindDataHelper
import  BaseData as BD
    
#======================================================================================================================
#======================================================================================================================
#功能:交易员信息
class Trader(object):
    def __init__(self, tradername, tradedate):
        self._name = tradername
        self._tradedate = tradedate
        self._hasOption = False         #该交易员是否交易期权,默认为False
        
        self._productmap ={}        #每个交易员，按品种汇总  key = 品种代码，   value = [持仓盈亏, 成交盈亏,手续费 ]
        self._codemappl = {}        #每个交易员， 按代码汇总  key = 代码        value = [持仓盈亏, 成交盈亏,手续费 ]
        
        self._realaccount = tradername
        pass
    
    #外部调用，设置外部数据来源
    def setBaseData(self,  BaseCtpData, ProductData ):
        self._baseCtpData = BaseCtpData
        self._productData = ProductData
        pass
    
    #计算期权
    def parseClearData2(self, prefilepath, filepath, caltype, sub_product):
        if caltype == 'ctp1':
            self.calcOptionData(prefilepath, filepath, self._tradedate,  sub_product )
        pass

    def parseClearData(self, filepath, caltype):
        if caltype == 'jn':
            self.parseJNClearData(filepath)
        elif caltype == 'ctp2':
            self.parseCtpClearData(filepath)
        pass

    #==============================================================================================
    #以下计算金牛 场内期货的盈亏
    #==============================================================================================
    def parseJNFile_2(self, filepath):      #第二版
        fr = codecs.open(filepath, 'r','utf-8')
        lines = [l.rstrip('\n') for l in fr.readlines()]
        if len(lines) == 0:
            return 
    
        flg = 0         #用于标识 当前需要进行何种操作
        tmap = []       #用于存储 成交明细里面的每一个字段
        pmap = []       #用于存储 持仓汇总里面的每一个字段
        
        for liner in lines:     
            line = liner.strip()
            if line.startswith('日期：'):
                self._realdate = line[line.find('：')+1:]
            elif line.startswith('账户：'):
                self._realaccount = line[line.find('：')+1:]
            elif line.strip() == '成交明细  Transaction Record' :
                flg = 10
            elif line.strip() == '持仓汇总 Positions' :
                flg = 20
            elif line.strip() == '成交明细':   #金牛20190613之前的格式
                flg = 30
            elif line.strip() == '持仓汇总':   #金牛20190613之前的格式
                flg = 40
            elif len(line) == 0:
                if (flg in range(10, 14)) or (flg in range(20, 24)) or (flg in range(30, 34)) or (flg in range(40, 44)):
                    pass
                else:
                    flg = 0
            elif re.match(r'^-*$',line):
                flg += 1
            else:
                if flg == 12:
                    args = [w.strip() for w in line.split('|')]
                    tmap.append(args) 
                elif flg == 22:
                    args = [w.strip() for w in line.split('|')]
                    pmap.append(args)
                elif flg == 32:
                    args = [w.strip() for w in line.split('|')]
                    tmap.append(args)
                elif flg == 42:
                    args = [w.strip() for w in line.split('|')]
                    pmap.append(args)
                pass
        
        #重整数据，主要是变换 数据类型
        #[1]成交明细
        if len(tmap) >0:
            tdf =pd.DataFrame(tmap)
            del tdf[0]    
            del tdf[11]    
            tdf[6] = tdf[6].astype('float')     #成交价
            tdf[9] = tdf[9].astype('float')     #手续费
            tdf[7] = tdf[7].astype('int')       #手数
            
            tdf[11] = tdf[4].apply(lambda bs : -1 if bs == '买' else 1)
        else:
            tdf = None
        pass
    
        #[2]持仓汇总
        if len(pmap) > 0:
            pdf =pd.DataFrame(pmap)
            del pdf[0]    
            del pdf[9] 
            pdf[3] = pdf[3].astype('int')           #持仓量
                  
            pdf[9] = pdf[2].apply(lambda bs : 1 if bs == '买' else -1)
        else:
            pdf = None
        
        return  (tdf, pdf)
    
    #外部调用，分析金牛结算单，并计算
    def parseJNClearData(self,  filepath):
        tdf, pdf= self.parseJNFile_2(filepath)
        
        #计算手续费，成交盈亏
        def lambda_shouxufei(x):
            code = x[3]
           
            product = self._productData.getProductByCode(code)
            if product in self._productmap:
                pass
            else:
                self._productmap[product] = [0.0, 0.0, 0.0]
                pass
            
            if code in self._codemappl:
                pass
            else:
                self._codemappl[code] = [0.0, 0.0, 0.0]
           
            mo_vm = self._productData.getMultiple(code)
            precloseprice,_  =  self._baseCtpData.getClosePrice(code)
            
            fee = x[9]
            single = mo_vm * (x[6] - precloseprice) * x[7] * x[11]
            
            self._productmap[product][2] += fee
            self._productmap[product][1] += single
            
            self._codemappl[code][2] += fee
            self._codemappl[code][1] += single
           
            return single
        
        
        self._future_fee = 0       #手续费
        self._future_vol = 0                #成交量
        pcpl = 0.0
        if tdf is not None:
            if len(tdf) >0:
                self._future_fee = round(tdf[9].sum(),2)       #手续费
                self._future_vol = tdf[7].sum()                #成交量
                pcpl = round(tdf.apply( lambda_shouxufei, axis = 1).sum(), 2)
            
        #计算持仓盈亏
        def lambda_ccpl(x):
            code = x[1]
            
            mo_vm = self._productData.getMultiple(code)
            precloseprice,closeprice  =  self._baseCtpData.getClosePrice(code)
            
            product = self._productData.getProductByCode(code)
            if product in self._productmap:
                pass
            else:
                self._productmap[product] = [0.0, 0.0, 0.0]
                pass
            
            if code in self._codemappl:
                pass
            else:
                self._codemappl[code] = [0.0, 0.0, 0.0]
                
            single = mo_vm * x[9] * x[3] * (closeprice - precloseprice)
        
            self._productmap[product][0] += single
            self._codemappl[code][0] += single
            return single
        
        ccpl = 0.0
        if pdf is not None:
            if len(pdf) > 0 :
                ccpl = round(pdf.apply( lambda_ccpl, axis = 1).sum(),2)

        self._future_ccpl = ccpl               #持仓盈亏
        self._future_pcpl = pcpl               #平仓盈亏
    
    #==============================================================================================
    #以下计算CTP 场内期货的盈亏
    #==============================================================================================
    def parseCtpFile_2(self, filepath):      #第二版
        fr = codecs.open(filepath, 'r','utf-8')
        lines = [l.rstrip('\n') for l in fr.readlines()]
        if len(lines) == 0:
            return None
        
        flg = 0
        tmap = []
        pmap = []
       
        for liner in lines:     
            line = liner.strip()

            if line.strip() == '成交记录 Transaction Record' :
                flg = 10      
            elif line.strip() == '持仓汇总 Positions' :
                flg = 20
            elif len(line) == 0:
                if 10 <= flg <= 12:
                    pass
                elif 20 <= flg <= 22:
                    pass
                else:
                    flg = 0
            elif re.match(r'^-*$',line):
                flg += 1
            else:
                if flg == 12:
                    args = [w.strip() for w in line.split('|')]
                    tmap.append(args)  
                elif flg == 22:
                    args = [w.strip() for w in line.split('|')]
                    pmap.append(args)

        if len(tmap) >0:            
            tdf =pd.DataFrame(tmap)
            del tdf[0]    
            del tdf[15]    
           
            tdf[7] = tdf[7].astype('float') #成交价
            tdf[11] = tdf[11].astype('float')  #手续费
            tdf[8] = tdf[8].astype('int')  #手数
            
            tdf[15] = tdf[5].apply(lambda bs : -1 if bs == '买' else 1)
        else:
            tdf = None

        if len(pmap) >0:
            pdf =pd.DataFrame(pmap)
            del pdf[0]    
            del pdf[14] 
            pdf[3] = pdf[3].astype('int')   #买持仓量
            pdf[5] = pdf[5].astype('int')   #卖持仓量
        else:
            pdf = None

        return (tdf, pdf)
        
    def parseCtpClearData(self, filepath):
        tdf, pdf= self.parseCtpFile_2(filepath)
        
        
        #计算手续费，成交盈亏
        def lambda_shouxufei(x):
            code = x[4]
           
            product = self._productData.getProductByCode(code)
            if product in self._productmap:
                pass
            else:
                self._productmap[product] = [0.0, 0.0, 0.0]
                pass
            
            if code in self._codemappl:
                pass
            else:
                self._codemappl[code] = [0.0, 0.0, 0.0]
           
            mo_vm = self._productData.getMultiple(code)
            precloseprice,_  =  self._baseCtpData.getClosePrice(code)
            
            fee = x[11]
            single = mo_vm * (x[7] - precloseprice) * x[8] * x[15]
            
            self._productmap[product][2] += fee
            self._productmap[product][1] += single
            
            self._codemappl[code][2] += fee
            self._codemappl[code][1] += single
           
            return single
        
        
        self._future_fee = 0       #手续费
        self._future_vol = 0                #成交量
        pcpl = 0.0
        if tdf is not None:
            if len(tdf) >0:
                self._future_fee = round(tdf[11].sum(),2)       #手续费
                self._future_vol = tdf[8].sum()                #成交量
                pcpl = round(tdf.apply( lambda_shouxufei, axis = 1).sum(), 2)
            
        #计算持仓盈亏
        def lambda_ccpl(x):
            code = x[2]
            
            mo_vm = self._productData.getMultiple(code)
            precloseprice,closeprice  =  self._baseCtpData.getClosePrice(code)
            
            product = self._productData.getProductByCode(code)
            if product in self._productmap:
                pass
            else:
                self._productmap[product] = [0.0, 0.0, 0.0]
                pass
            
            if code in self._codemappl:
                pass
            else:
                self._codemappl[code] = [0.0, 0.0, 0.0]
                
            vol = x[3] - x[5]       #买为1， 卖为-1
            single = mo_vm * vol * (closeprice - precloseprice)
        
            self._productmap[product][0] += single
            self._codemappl[code][0] += single
            return single
        
        ccpl = 0.0
        if pdf is not None:
            if len(pdf) > 0 :
                ccpl = round(pdf.apply( lambda_ccpl, axis = 1).sum(),2)

        self._future_ccpl = ccpl               #持仓盈亏
        self._future_pcpl = pcpl               #平仓盈亏
        pass

    #==============================================================================================
    #以下计算CTP 场内期权的盈亏
    #==============================================================================================
    def helper_prePosotion(self, prectpfile, dt_today, sub_product):
        fr = codecs.open(prectpfile, 'r','utf-8')
        lines = [l.rstrip('\n') for l in fr.readlines()]
        if len(lines) == 0:
            return None
        
        flg = 0
        pmap = []
       
        for liner in lines:     
            line = liner.strip()
            
            if line.strip() == '持仓汇总 Positions' :
                flg = 20
            elif len(line) == 0:
                if flg in range(20, 24):
                    pass
                else:
                    flg = 0
            elif re.match(r'^-*$',line):
                flg += 1
            else:
                if flg == 22:
                    args = [w.strip() for w in line.split('|')]
                    pmap.append(args)
        
        if len(pmap) >0:
            pdf =pd.DataFrame(pmap)
            del pdf[0]    
            del pdf[14] 
            pdf[3] = pdf[3].astype('int')   #买持仓量
            pdf[5] = pdf[5].astype('int')   #卖持仓量
        else:
            return None
    
        if len(pdf) == 0:
            return None
        
        if sub_product == 'CU':
            mop = pdf[[bool(re.match(r'^cu\d+[CP]\d+$',i)) for i in pdf[2]]]  
        elif sub_product == 'RU':
            mop = pdf[[bool(re.match(r'^ru\d+[CP]\d+$',i)) for i in pdf[2]]]  
        elif sub_product == 'CF':
            mop = pdf[[bool(re.match(r'^CF\d+[CP]\d+$',i)) for i in pdf[2]]] 
        elif sub_product == 'i':
            mop = pdf[[bool(re.match(r'^i\d+-[CP]-\d+$',i)) for i in pdf[2]]] 

        if len(mop) ==0:
            return None
        
        #准备计算 昨日持仓汇总
        def lambda_ccpl(x):
            code = x[2]
            exch_id = self._productData.getExchange(code)
            mo_vm = self._productData.getMultiple(code)
            precloseprice,closeprice  =  self._baseCtpData.getClosePrice(code)
            
            bret = self._wind.isOptionLastdate(code, dt_today, exch_id) 
            print("==============> code = ", code, dt_today, bret)
            #如果当日是场内某期权的最后到期日，则置为0
            if bret:
                closeprice =0.0
            
            vol = x[3] - x[5]       #买为1， 卖为-1
            single = mo_vm * vol * (closeprice - precloseprice)
            return single
        
        ccpl = round(mop.apply(lambda_ccpl, axis = 1).sum(),2)
        
        return ccpl
    
    def helper_todayTrade(self, ctpfile, dt_today, sub_product):
        fr = codecs.open(ctpfile, 'r','utf-8')
        lines = [l.rstrip('\n') for l in fr.readlines()]
        if len(lines) == 0:
            return None
        
        flg = 0
        tmap = []
   
        for liner in lines:     
            line = liner.strip()
                         
            if line.strip() == '成交记录 Transaction Record' :
                flg = 10
            elif len(line) == 0:
                if flg in range(10, 14):
                    pass
                else:
                    flg = 0
            elif re.match(r'^-*$',line):
                flg += 1
            else:
                if flg == 12:
                    args = [w.strip() for w in line.split('|')]
                    tmap.append(args)            
                
        if len(tmap) >0:            
            tdf =pd.DataFrame(tmap)
            del tdf[0]    
            del tdf[15]    
           
            tdf[7] = tdf[7].astype('float') #成交价
            tdf[11] = tdf[11].astype('float')  #手续费
            tdf[8] = tdf[8].astype('int')  #手数
            
            tdf[15] = tdf[5].apply(lambda bs : 1 if bs == '买' else -1)
        else:
            return None

        if len(tdf) == 0:
            return None
        
        def lambda_shouxufei(x):
            code = x[4]
            exch_id = self._productData.getExchange(code)
            mo_vm = self._productData.getMultiple(code)
            _,closeprice  =  self._baseCtpData.getClosePrice(code)
            
            bret = self._wind.isOptionLastdate(code, dt_today, exch_id)   #如果当日是场内某期权的最后到期日，则置为0
            if bret:
                closeprice =0.0
            
            single = mo_vm * ( closeprice - x[7] ) * x[8] * x[15]
            
            return single
        
        if sub_product == 'CU':
            mot = tdf[[bool(re.match(r'^cu\d+[CP]\d+$',i)) for i in tdf[4]]]
        elif sub_product == 'RU':
            mot = tdf[[bool(re.match(r'^ru\d+[CP]\d+$',i)) for i in tdf[4]]]
        elif sub_product == 'CF':
            mot = tdf[[bool(re.match(r'^CF\d+[CP]\d+$',i)) for i in tdf[4]]] 
        elif sub_product == 'i':
            mot = tdf[[bool(re.match(r'^i\d+-[CP]-\d+$',i)) for i in tdf[4]]] 
            
        if len(mot) > 0:
            self._option_fee = round(mot[11].sum(),2) 
            self._option_vol = mot[8].sum() 
            pcpl = round( mot.apply( lambda_shouxufei, axis = 1).sum(), 2)
        else:
            self._option_fee = 0
            self._option_vol = 0
            pcpl =0
            pass
        
        return pcpl
    
    def calcOptionData(self, prectpfile,  ctpfile, dt_today, sub_product):
        
        self._hasOption = True
        
        self._wind = WindDataHelper.WindCUOptionData()
        self._wind.start()
        
        #[1]获取昨日持仓，计算持仓盈亏
        ccpl = self.helper_prePosotion(prectpfile, dt_today, sub_product)
        if ccpl is None:
            self._option_ccpl = 0.0
        else:
            self._option_ccpl = ccpl
        
        #[2]计算今日成交盈亏
        pcpl = self.helper_todayTrade(ctpfile, dt_today, sub_product)
        if pcpl is None:
            self._option_pcpl =0.0
        else:
            self._option_pcpl = pcpl
            
        pass
    
    
    
    #==============================================================================================
    #以下输出结果
    #==============================================================================================
    #外部调用，输出结果
    def printself(self):
        print('[姓名]', self._name)
        if self._hasOption:
            print('[期权成交量]', self._option_vol)
            print('[期权手续费]', self._option_fee)
            print('[期权持仓盈亏]', self._option_ccpl)
            print('[期权平仓盈亏]', self._option_pcpl)
            print('[期权当日盈亏]', self._option_ccpl + self._option_pcpl)
            print('[期权当日盈亏(扣手续费)]', self._option_ccpl + self._option_pcpl - self._option_fee)
        else:
            print('[成交量]', self._future_vol)
            print('[手续费]', self._future_fee)
            print('[持仓盈亏]', self._future_ccpl)
            print('[平仓盈亏]', self._future_pcpl)
            print('[当日盈亏]', self._future_ccpl + self._future_pcpl)
            print('[当日盈亏(扣手续费)]', self._future_ccpl + self._future_pcpl - self._future_fee)
        pass
    
    #外部调用，以list形式返回期货的结果
    def getFutureResult(self):
        if self._hasOption:
            return None

        li = []
        li.append(self._realaccount)
        li.append(self._future_pcpl)
        li.append(self._future_ccpl)
        li.append(self._future_ccpl + self._future_pcpl)
        li.append(self._future_fee)
        li.append(self._future_ccpl + self._future_pcpl - self._future_fee)
        return li
    
    #外部调用，以list形式返回期权的结果
    def getOptionResult(self):
        if self._hasOption:
            li = []
            li.append(self._realaccount)
            li.append(self._option_pcpl)
            li.append(self._option_ccpl)
            li.append(self._option_ccpl + self._option_pcpl)
            li.append(self._option_fee)
            li.append(self._option_ccpl + self._option_pcpl - self._option_fee)
            return li
        
        return None
    
    def getProductTotal(self):
        return self._name, self._productmap

    def getCodeTotal(self):
        return self._name, self._codemappl

#======================================================================================================================
#======================================================================================================================
def make_Total(traderlist,  workbook,  worksheet):
    # 设置C到H单元格宽度15像素
    cols = [ 'C', 'D', 'E', 'F', 'G', 'H']
    for col in cols:
        worksheet.column_dimensions[ col ].width = 15
    
    
    rows = []
    rows.append( [' ', '平仓盈亏', '持仓盯市盈亏', '盈亏', '手续费', '扣除手续费后']   )
    
    for i in traderlist:
        ret = i.getFutureResult()
        if ret is None:
            pass
        else:
            rows.append( ret )
        
    for i in traderlist:
        ret = i.getOptionResult()
        if ret is None:
            pass
        else:
            rows.append(ret)
    
    ft = Font(name=u'宋体',size=11)
    ft_bold = Font(name=u'宋体',size=11, bold = True)
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)
       
    i =0
    for row in rows:
        j = 0
        for col in cols:
            cell = worksheet.cell(i+1, j+3)
            cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平居中，垂直居中
            cell.value = row[j]
            #print(row[j])
            cell.border = border
            
            if i == 0:
                cell.font = ft_bold
            else:
                cell.font = ft
                
                
                if j >0:
                    cell.number_format = '#,##0.00_);[red](#,##0.00)'
            
            
            j += 1
            
        i += 1
        pass
    
    j =0
    for col in cols:
        cell = worksheet.cell(i+1, j+3)
        if j ==0:
            cell.value = "合计"
            pass
        else:
            cell.value = '=SUM({0}{1}:{2}{3})'.format(col,2, col, i)
            cell.number_format = '#,##0.00_);[red](#,##0.00)'
            pass
        j += 1
        pass
    
    pass

def make_EveryTrader(traderlist,  workbook,  worksheet):
    
    rowno = 12
    worksheet.cell(rowno, 3).value = "持仓盈亏"
    worksheet.cell(rowno, 4).value = "平仓盈亏"
    worksheet.cell(rowno, 5).value = "盈亏"
    worksheet.cell(rowno, 6).value = "手续费"
    for trader in traderlist:
        tradername, promap = trader.getProductTotal()
        
        if promap is None:
            continue
        
        lines = len(promap)
        if lines ==0:
            continue
        
        worksheet.cell(rowno+1, 1).value = tradername
        for k, v in promap.items():
            worksheet.cell(rowno+1, 2).value = k
            worksheet.cell(rowno+1, 3).value = v[0]
            worksheet.cell(rowno+1, 4).value = v[1]
            worksheet.cell(rowno+1, 5).value = v[0] + v[1]
            worksheet.cell(rowno+1, 6).value = v[2]
            
            rowno += 1
            pass
    
    
    rowno += 4
    for trader in traderlist:
        tradername, codemap = trader.getCodeTotal()
        
        if codemap is None:
            continue
        
        lines = len(codemap)
        if lines ==0:
            continue
        
        worksheet.cell(rowno+1, 1).value = tradername
        for k, v in codemap.items():
            worksheet.cell(rowno+1, 2).value = k
            worksheet.cell(rowno+1, 3).value = v[0]
            worksheet.cell(rowno+1, 4).value = v[1]
            worksheet.cell(rowno+1, 5).value = v[0] + v[1]
            worksheet.cell(rowno+1, 6).value = v[2]
            
            rowno += 1
            pass
    
    pass

    



def makeReport(rootdir, traderlist, date_dt):
    reportdir = os.path.join(rootdir, "report")
    
    filename = "{0}-{1}-{2}损益日报表(收盘价).xlsm".format( date_dt.year,  date_dt.month, date_dt.day)
    filepath = os.path.join(reportdir, filename)
    
    datadir = os.path.join(rootdir, "dataotc")
    srcpath = os.path.join(  datadir, "Blank模板勿删(收盘价).xlsm")
    if os.path.exists(filepath):
        pass
    else:
        shutil.copyfile(srcpath,   filepath)
    
    workbook = openpyxl.load_workbook(filepath, keep_vba = True)
    #按序号获取sheet
    worksheet = workbook.worksheets[2]
    
    #填写汇总盈亏
    make_Total(traderlist, workbook, worksheet)
    
    #填写按交易员按品种汇总
    make_EveryTrader(traderlist, workbook, worksheet)
    
    workbook.save(filepath)
    workbook.close()
    pass


#功能: 扫描结算数据目录，寻找某个账号在dt_today的文件
def scan_find(rootdir, dt_today, account):
    curdir = os.path.join( rootdir, "datactp_jn")
    if not os.path.exists(curdir):
        return None

    date_str1 = "%04d-%02d-%02d" % ( dt_today.year,  dt_today.month, dt_today.day )
    date_str2 = str(  dt_today.year*10000 + dt_today.month*100 +dt_today.day ) 

    oblist = os.listdir(curdir)
    for one in oblist:
        opath = os.path.join(curdir, one)
        
        if os.path.isfile(opath):
            if os.path.splitext(one)[1] == '.txt':
                if one.find(date_str1) >= 0 or one.find(date_str2) >= 0:
                    if one.startswith( account ):
                        return opath
    
    return None

#======================================================================================================================
#======================================================================================================================
def main(rootdir, dt_today):
    tmap = {    
                "林秉玮":('0009', 'jn'), 
                "谢晨星":('0019', 'jn'), 
                "何剑桥":('0029', 'jn'),  
                "刘剑溥":('0039', 'jn'),
                "场内CU"  :('8001888888', 'ctp1', 'CU'),   
                "场内RU"  :('8001888888', 'ctp1', 'RU'),
                "场内CF"  :('8001888888', 'ctp1', 'CF'),
                "场内i" :('8001888888', 'ctp1', 'i'),
                "911005":('9110000005', 'ctp2'),
                "911002":('9110000002', 'ctp2')
            }

    #[1]基础数据
    oproduct = BD.CProductData()
    oholiday = BD.CHolidayData()
    pre_workday = oholiday.preWorkDay(dt_today)

    baseobj = BD.CSnapData4(dt_today,  pre_workday)
    baseobj.start()
    
    traderlist = []
    for k, v in tmap.items():
        if v[1] == 'ctp1':  #计算ctp场内期权，需要2日的结算单
            prefilepath = scan_find(rootdir, pre_workday, v[0] )
            filepath = scan_find(rootdir, dt_today, v[0] )
            if prefilepath is None:
                continue
            if filepath is None:
                continue
            pass
        else:  #计算金牛，或ctp, 需要当日的结算单
            filepath = scan_find(rootdir, dt_today, v[0] )
            if filepath is None:
                continue

        #every trader
        trader = Trader(k, dt_today)
        trader.setBaseData(baseobj, oproduct )
        if v[1] == 'ctp1':
            trader.parseClearData2(prefilepath, filepath, v[1], v[2] )
        else:
            trader.parseClearData(filepath, v[1] )

        trader.printself()
        traderlist.append( trader )
        pass

    #[4]汇总并输出报表
    makeReport(rootdir, traderlist, dt_today)
    
    pass








if __name__ == '__main__':
    print("main end!!!!!!!!")